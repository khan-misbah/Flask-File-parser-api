from flask import Flask,jsonify, request
from flask_sqlalchemy import SQLAlchemy
from flask_marshmallow import Marshmallow
from openpyxl  import load_workbook
import os
import sqlite3
from werkzeug.utils import secure_filename



app = Flask(__name__)

app.config['SQLALCHEMY_DATABASE_URI'] = "sqlite:///student.db"
app.config['SQLALCHEMY_TRACK_MODIFICATIONS']= False

db = SQLAlchemy(app)
ma = Marshmallow(app)
# obj = user_model

class Student(db.Model):
    id=db.Column(db.Integer, primary_key=True)
    name=db.Column(db.String(200),nullable=False)
    age=db.Column(db.Integer)
    gender=db.Column(db.String(20))
    phone=db.Column(db.Integer)
    city=db.Column(db.String(50))
    def __repr__(self) -> str:
        return f"{self.id} - {self.name}"

class StudentSchema(ma.Schema):
    class Meta:
        fields=["name","Age","Gender", "Phone","City"]
   
student_schema=StudentSchema()
students_schema=StudentSchema(many=True)

@app.route('/student/<id>') 
def get_studnet(id):
    stud = Student.query.filter_by(id=id).first()
    if stud:
        return student_schema.jsonify(stud)
        
    return jsonify({'message':'student name not found'})


# @app.route('/students',methods=['GET']) 
# def get_students():
#     stud=Student.query.all()
#     if stud:
#         return students_schema.jsonify(stud)
        
#     return jsonify({'message':'student not found'})




@app.route('/student',methods=['POST','GET'])
def create_student():
    if request.method=="POST":
        data= request.get_json() 
        print(data,"............")
        name=data["name"]
        age=data["age"]
        gender=data["gender"]
        phone=data["phone"]
        city=data["city"]
        print("name..........",name)
        print("Age........",age)
        studs=Student(name=name,age=age,gender=gender,phone= phone,city=city)
        db.session.add(studs)
        db.session.commit()
    
    studs=Student.query.all()
    print(studs)
    # return jsonify("add value")
    return students_schema.jsonify(studs)


@app.route("/student/addmultiple", methods=["POST"])

def user_add_multiple_data():
    data=request.get_json()
    print("this is data",data)
    for userdata in data:
        print(userdata["name"],"......")
        
        name=userdata["name"]
        age=userdata["age"]
        gender=userdata["gender"]
        phone=userdata["phone"]
        city=userdata["city"]

        main = Student(name=name,age=age,gender=gender,phone=phone,city=city)
        db.session.add(main)
        db.session.commit()
        main=Student.query.all()
 
    return students_schema.jsonify(main)


@app.route("/student/uploadFile", methods=["POST"])
def uploadFile():
    file=request.files["filename"]
    print("file_data.......",file)
    # print("filename.......",file_data)
    # file.save(os.path.join(app.config['file_name'], file))

    file.save(os.path.join('static',secure_filename(file.filename)))
    wb = load_workbook(file) 
    sheet = wb.active
    conn = sqlite3.connect('mydata.db')


    cursor = conn.cursor()

    cursor.execute('''CREATE TABLE IF NOT EXISTS data (name TEXT, age INTEGER, gender TEXT, phone INTEGER,city TEXT)''')
    for row in sheet.iter_rows(min_row=5):
        name = row[0].value
        age = row[1].value
        gender = row[2].value
        phone = row[3].value
        city = row[4].value


        cursor.execute('''INSERT OR IGNORE INTO data(name,age,gender,phone,city) VALUES (?, ?,?,?,?)''',(name,age,gender,phone,city))

        conn.commit()
        conn.close()
        return jsonify({"massage": "Upload File Successfully"})

        # return jsonify({'message': 'addded in db'})



    # wb_obj = openpyxl.load_workbook(file)
 
    # sheet_obj = wb_obj.active
 
    # max_col = sheet_obj.max_column
    # max_row = sheet_obj.max_rows

    # for i in range(1, max_col + 1):
    #  cell_obj = sheet_obj.cell(row = 2, column = i)
    #  for i in (0,max_row+1):
    #      cell_obj= sheet_obj.cell(row= 2, rows = i)
    # print(cell_obj.value,  " ...///////")
    # # # for i in range(1, dataframe1 + 1):
    #    cell_obj = dataframe.cell(row = 2, column = i)
    # print(cell_obj.value, end = " ")
    # max_col = dataframe1.max_column
#    for i in range(1, max_row + 1):
    #  /   cell_obj = dataframe1.cell(row = 2, column = i)
        # print(cell_obj.value,  ".....")
    # for row in range(0,dataframe1.max_row):
        # for col in dataframe1.iter_cols(1,dataframe1.max_column):
            # print(col[row].value,".........)

 # path = (file)
    # wb_obj = openpyxl.load_workbook(file)

    # sheet_obj = wb_obj.active
    # max_col = sheet_obj.max_column
    # for i in range(1, max_col + 1):
    #     cell_obj = sheket_obj.cell(row = 2, column = i)
    # print(cell_obj.value, end = " ")

    # main=Student.query.all()

    # # return students_schema.jsonify(file)







@app.route('/')
def home():
    return "hey"


if __name__=="__main__":
    app.run(debug=True)

with app.app_context():
    db.create_all()